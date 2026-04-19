from __future__ import annotations

import base64
from io import BytesIO
from pathlib import Path
import tempfile
import unittest

from openpyxl import Workbook

from app.schemas import (
    ClinicalDataBundle,
    ClinicalDataCsvBundle,
    ClinicalDataCsvImportRequest,
    ClinicalDataExcelBundle,
    ClinicalDataExcelImportRequest,
    ClinicalDataImportRequest,
    DiagnosisRecord,
    ExternalClinicalMetadata,
    ImportedClinicalPatient,
    LabRecord,
    PatientOverview,
    VitalRecord,
)
from app.services import import_service
from app.services.diagnosis_service import get_diagnoses
from app.services.lab_service import get_recent_labs
from app.services.patient_service import get_all_hadm_ids, get_patient_overview
from app.services.vital_service import get_recent_vitals


class ExternalClinicalImportIntegrationTest(unittest.TestCase):
    def setUp(self) -> None:
        self.tempdir = tempfile.TemporaryDirectory()
        self.original_import_dir = import_service.EXTERNAL_CLINICAL_IMPORT_DIR
        import_service.EXTERNAL_CLINICAL_IMPORT_DIR = Path(self.tempdir.name)
        import_service.reset_external_import_cache()

    def tearDown(self) -> None:
        import_service.EXTERNAL_CLINICAL_IMPORT_DIR = self.original_import_dir
        import_service.reset_external_import_cache()
        self.tempdir.cleanup()

    def test_imported_patient_flows_through_existing_queries(self) -> None:
        payload = ClinicalDataImportRequest(
            dataset_name="unittest-demo",
            bundle=ClinicalDataBundle(
                metadata=ExternalClinicalMetadata(
                    name="Unit Test Demo",
                    source="unittest",
                ),
                patients=[
                    ImportedClinicalPatient(
                        hadm_id=900001,
                        patient_overview=PatientOverview(
                            subject_id=500001,
                            gender="F",
                            age=67,
                            admittime="2026-04-16T08:30:00Z",
                            admission_type="URGENT",
                        ),
                        diagnoses=[
                            DiagnosisRecord(
                                hadm_id=900001,
                                seq_num=1,
                                icd_code="A41.9",
                                icd_version=10,
                            )
                        ],
                        labs=[
                            LabRecord(
                                hadm_id=900001,
                                itemid=50813,
                                label="Lactate",
                                charttime="2026-04-16T09:00:00Z",
                                value="3.2",
                                valuenum=3.2,
                                valueuom="mmol/L",
                                flag="abnormal",
                            )
                        ],
                        vitals=[
                            VitalRecord(
                                hadm_id=900001,
                                itemid=220045,
                                label="Heart Rate",
                                charttime="2026-04-16T09:05:00Z",
                                value="112",
                                valuenum=112,
                                valueuom="bpm",
                                warning=1,
                            )
                        ],
                    )
                ],
            ),
        )

        summary = import_service.import_clinical_data(payload)

        self.assertEqual(summary["patient_count"], 1)
        self.assertEqual(summary["hadm_ids"], [900001])
        self.assertEqual(summary["record_counts"]["diagnoses"], 1)
        self.assertEqual(summary["record_counts"]["labs"], 1)
        self.assertEqual(summary["record_counts"]["vitals"], 1)

        import_service.reset_external_import_cache()

        self.assertIn(900001, get_all_hadm_ids())
        self.assertEqual(get_patient_overview(900001)["age"], 67)
        self.assertEqual(get_diagnoses(900001)[0]["icd_code"], "A41.9")
        self.assertEqual(get_recent_labs(900001, "lactate")[0]["label"], "Lactate")
        self.assertEqual(get_recent_vitals(900001, "heart")[0]["label"], "Heart Rate")

    def test_csv_import_flows_through_existing_queries(self) -> None:
        payload = ClinicalDataCsvImportRequest(
            dataset_name="csv-demo",
            csv_bundle=ClinicalDataCsvBundle(
                patients_csv=(
                    "hadm_id,subject_id,gender,age,admittime,admission_type\n"
                    "900002,500002,M,54,2026-04-17T08:00:00Z,EMERGENCY\n"
                ),
                diagnoses_csv=(
                    "hadm_id,seq_num,icd_code,icd_version\n"
                    "900002,1,J18.9,10\n"
                ),
                labs_csv=(
                    "hadm_id,itemid,label,charttime,value,valuenum,valueuom,flag\n"
                    "900002,50912,Creatinine,2026-04-17T09:00:00Z,1.8,1.8,mg/dL,abnormal\n"
                ),
                vitals_csv=(
                    "hadm_id,itemid,label,charttime,value,valuenum,valueuom,warning\n"
                    "900002,220045,Heart Rate,2026-04-17T09:05:00Z,108,108,bpm,1\n"
                ),
            ),
        )

        summary = import_service.import_clinical_csv_data(payload)

        self.assertEqual(summary["patient_count"], 1)
        self.assertEqual(summary["hadm_ids"], [900002])

        import_service.reset_external_import_cache()

        self.assertIn(900002, get_all_hadm_ids())
        self.assertEqual(get_patient_overview(900002)["gender"], "M")
        self.assertEqual(get_diagnoses(900002)[0]["icd_code"], "J18.9")
        self.assertEqual(get_recent_labs(900002, "creatinine")[0]["label"], "Creatinine")
        self.assertEqual(get_recent_vitals(900002, "heart")[0]["label"], "Heart Rate")

    def test_import_history_and_delete(self) -> None:
        first = import_service.import_clinical_data(
            ClinicalDataImportRequest(
                dataset_name="first-demo",
                bundle=ClinicalDataBundle(
                    metadata=ExternalClinicalMetadata(name="first-demo"),
                    patients=[
                        ImportedClinicalPatient(
                            hadm_id=910001,
                            patient_overview=PatientOverview(subject_id=510001),
                        )
                    ],
                ),
            )
        )
        second = import_service.import_clinical_data(
            ClinicalDataImportRequest(
                dataset_name="second-demo",
                bundle=ClinicalDataBundle(
                    metadata=ExternalClinicalMetadata(name="second-demo"),
                    patients=[
                        ImportedClinicalPatient(
                            hadm_id=910002,
                            patient_overview=PatientOverview(subject_id=510002),
                        )
                    ],
                ),
            )
        )

        history = import_service.list_imported_datasets()

        self.assertEqual(len(history), 2)
        self.assertEqual(
            {item["import_id"] for item in history},
            {first["import_id"], second["import_id"]},
        )
        self.assertNotEqual(first["import_id"], second["import_id"])

        deleted = import_service.delete_imported_dataset(second["import_id"])

        self.assertEqual(deleted["import_id"], second["import_id"])
        self.assertNotIn(910002, get_all_hadm_ids())
        self.assertIn(910001, get_all_hadm_ids())

    def test_excel_import_flows_through_existing_queries(self) -> None:
        workbook = Workbook()
        patients_sheet = workbook.active
        patients_sheet.title = "Patients"
        patients_sheet.append(["hadm_id", "subject_id", "gender", "age", "admittime"])
        patients_sheet.append([900003, 500003, "F", 61, "2026-04-18T08:00:00Z"])

        diagnoses_sheet = workbook.create_sheet("Diagnoses")
        diagnoses_sheet.append(["hadm_id", "seq_num", "icd_code", "icd_version"])
        diagnoses_sheet.append([900003, 1, "I10", 10])

        labs_sheet = workbook.create_sheet("Labs")
        labs_sheet.append(
            ["hadm_id", "itemid", "label", "charttime", "value", "valuenum", "valueuom", "flag"]
        )
        labs_sheet.append([900003, 50912, "Creatinine", "2026-04-18T09:00:00Z", "1.4", 1.4, "mg/dL", ""])

        vitals_sheet = workbook.create_sheet("Vitals")
        vitals_sheet.append(
            ["hadm_id", "itemid", "label", "charttime", "value", "valuenum", "valueuom", "warning"]
        )
        vitals_sheet.append([900003, 220045, "Heart Rate", "2026-04-18T09:05:00Z", "104", 104, "bpm", 0])

        buffer = BytesIO()
        workbook.save(buffer)
        workbook_base64 = base64.b64encode(buffer.getvalue()).decode("utf-8")

        payload = ClinicalDataExcelImportRequest(
            dataset_name="excel-demo",
            excel_bundle=ClinicalDataExcelBundle(
                workbook_base64=workbook_base64,
                workbook_name="excel-demo.xlsx",
            ),
        )

        summary = import_service.import_clinical_excel_data(payload)

        self.assertEqual(summary["patient_count"], 1)
        self.assertEqual(summary["hadm_ids"], [900003])

        import_service.reset_external_import_cache()

        self.assertIn(900003, get_all_hadm_ids())
        self.assertEqual(get_patient_overview(900003)["age"], 61)
        self.assertEqual(get_diagnoses(900003)[0]["icd_code"], "I10")
        self.assertEqual(get_recent_labs(900003, "creatinine")[0]["label"], "Creatinine")
        self.assertEqual(get_recent_vitals(900003, "heart")[0]["label"], "Heart Rate")


if __name__ == "__main__":
    unittest.main()
